#!/usr/bin/env python3
"""
微信批量发送助手 — CLI 工具（macOS / Windows）
GUI 增强版：支持 --json 输出
"""
from __future__ import annotations

# 确保用户 site-packages 在 import 路径最前面（打包后系统 Python 也能找到 openpyxl 等依赖）
import sys as _sys
try:
    import site as _site
    _site.addusersitepackages(_sys.path)
except Exception:
    pass

import argparse
import json
import os
import platform
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional


def _real_home() -> Path:
    """返回真实的用户 home 目录，绕过 Python pkg 改写的 $HOME 环境变量。"""
    if sys.platform == "win32":
        return Path(os.environ.get("USERPROFILE", os.environ.get("HOMEDRIVE", "C:\\") + "\\Users\\" + os.environ.get("USERNAME", "")))
    return Path(os.path.expanduser("~"))

try:
    import openpyxl
    from rich.console import Console
    from rich.table import Table
    from rich.prompt import Confirm, Prompt
    from rich.panel import Panel
    console = Console()
    HAS_RICH = True
except ImportError:
    HAS_RICH = False
    class _FakeConsole:
        def print(self, *a, **kw): print(*a)
    console = _FakeConsole()

IS_MAC = platform.system() == "Darwin"
IS_WINDOWS = platform.system() == "Windows"

# 脚本自身所在目录（兼容 dev 和打包后路径）
_SELF_DIR   = Path(__file__).resolve().parents[0]          # .../python/app/
_SCRIPTS_DIR = _SELF_DIR.parent / "scripts"                  # .../python/scripts/
CFG_PATH   = _real_home() / ".wechat-sender" / "config.json"
APPLE_SCRIPT = _SCRIPTS_DIR / "wechat_send_mac.applescript"
SHEET_TASKS = "发送任务"
HEADER_ROW = 2

STATUS_WAITING = "待发送"
STATUS_RUNNING = "发送中"
STATUS_SUCCESS = "发送成功"
STATUS_FAILED = "发送失败"

COL = {
    "seq": "#",
    "app": "* 应用",
    "target": "* 联系人/群聊",
    "msg_type": "* 消息类型",
    "text": "* 文字内容",
    "image": "图片路径",
    "send_time": "发送时间",
    "repeat": "重复",
    "remark": "备注",
    "status": "状态",
}


@dataclass
class Task:
    row: int
    app: str
    target: str
    msg_type: str
    text: str
    image_path: str
    send_time: Optional[datetime]
    repeat: str
    status: str


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_cfg() -> dict:
    if not CFG_PATH.exists():
        return {}
    with open(CFG_PATH, "r", encoding="utf-8") as f:
        return json.load(f) or {}


def save_cfg(cfg: dict):
    CFG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CFG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def get_cfg():
    cfg = load_cfg()
    defaults = {
        "excel_path": "",
        "poll_seconds": 15,
        "dry_run": False,
        "send_interval": 5,
        "max_per_minute": 8,
    }
    for k, v in defaults.items():
        cfg.setdefault(k, v)
    return cfg


def find_columns(ws) -> dict:
    cols = {}
    for c in range(1, ws.max_column + 1):
        title = (ws.cell(HEADER_ROW, c).value or "").strip()
        if title:
            cols[title] = c
    return cols


def read_tasks(ws, cols) -> list[Task]:
    tasks: list[Task] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        app = str(ws.cell(r, cols.get(COL["app"], 0) or 1).value or "").strip()
        target = str(ws.cell(r, cols.get(COL["target"], 0) or 1).value or "").strip()
        msg_type = str(ws.cell(r, cols.get(COL["msg_type"], 0) or 1).value or "").strip()
        text = str(ws.cell(r, cols.get(COL["text"], 0) or 1).value or "").strip()
        image_path = str(ws.cell(r, cols.get(COL["image"], 0) or 1).value or "").strip()
        send_time_raw = ws.cell(r, cols.get(COL["send_time"], 0) or 1).value
        repeat_raw = ws.cell(r, cols.get(COL["repeat"], 0) or 1).value
        status = str(ws.cell(r, cols.get(COL["status"], 0) or 1).value or "").strip()
        if not any([app, target, msg_type, text, image_path, send_time_raw, repeat_raw, status]):
            continue
        task = Task(
            row=r, app=app, target=target, msg_type=msg_type,
            text=text, image_path=image_path,
            send_time=send_time_raw if isinstance(send_time_raw, datetime) else None,
            repeat=str(repeat_raw).strip() if repeat_raw else "",
            status=status,
        )
        tasks.append(task)
    return tasks


def set_status(ws, cols, row: int, status: str):
    ws.cell(row, cols[COL["status"]]).value = status


def validate_task(task: Task):
    # 当前仅支持微信自动化，忽略 app 字段
    if not task.target:
        raise ValueError("联系人/群聊不能为空")
    _valid_types = {"文字", "图片", "文字+图片"}
    if task.msg_type not in _valid_types:
        raise ValueError(f"不支持的消息类型: {task.msg_type}（仅支持：文字/图片/文字+图片）")
    if task.msg_type in {"文字", "文字+图片"} and not task.text:
        raise ValueError("文字内容不能为空")
    if task.msg_type in {"图片", "文字+图片"}:
        if not task.image_path:
            raise ValueError("图片消息缺少图片路径")
        if not Path(task.image_path).expanduser().exists():
            raise ValueError(f"图片不存在: {task.image_path}")


def should_send(task: Task, now: datetime) -> bool:
    if task.status.startswith(STATUS_SUCCESS) and not task.repeat:
        return False
    if task.send_time is None:
        return True
    return now >= task.send_time


def _ensure_accessibility():
    """触发 macOS Accessibility 权限对话框（自动弹出系统授权面板）。
    osascript 访问 System Events 时若未授权，系统自动弹出授权提示。"""
    if not IS_MAC:
        return
    try:
        subprocess.run([
            "osascript", "-e",
            "tell application \"System Events\"\n keystroke \"x\"\n end tell"
        ], capture_output=True, timeout=3)
    except Exception:
        pass


def call_sender(target: str, msg_type: str, text: str, image_path: str):
    img = str(Path(image_path).expanduser()) if image_path else ""
    if IS_MAC:
        _ensure_accessibility()
        cmd = ["osascript", str(APPLE_SCRIPT), target, msg_type, text, img]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if r.returncode != 0:
            err = r.stderr.strip() or "AppleScript 执行失败"
            print(f"[SEND_ERROR] {err}", file=sys.stderr)
            sys.exit(1)
        # 检查 stdout 中是否有 AppleScript error 关键字（script 内部 try 吞掉的错误）
        stdout_lower = r.stdout.lower()
        if "error" in stdout_lower and ("not found" in stdout_lower or "can’t" in stdout_lower or "permission" in stdout_lower):
            print(f"[SEND_ERROR] AppleScript 内部错误: {r.stdout.strip()}", file=sys.stderr)
            sys.exit(1)
    elif IS_WINDOWS:
        # 用 subprocess 调用（替代 importlib，避免 __import__  machinery 在打包环境报错）
        _win_script = _SCRIPTS_DIR / "wechat_send_win.py"
        cmd = [
            sys.executable,
            str(_win_script),
            "--call-single",
            "--target", target,
            "--msg-type", msg_type,
            "--text", text or "",
            "--image-path", img,
        ]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if r.returncode != 0:
            err = (r.stderr.strip() or r.stdout.strip() or "发送失败")
            print(f"[call_sender] {err}", file=sys.stderr)
            sys.exit(1)
    else:
        raise RuntimeError("当前系统不支持微信自动化（仅支持 macOS 和 Windows）")


# ─── 命令 ────────────────────────────────────────────────

def cmd_config_show(args):
    cfg = get_cfg()
    # JSON 模式（供 GUI 调用）
    if getattr(args, 'json', False):
        print(json.dumps(cfg, ensure_ascii=False))
        return

    os_label = "🍎 macOS" if IS_MAC else ("🪟 Windows" if IS_WINDOWS else "❓ Unknown")
    if HAS_RICH:
        table = Table(title=f"📋 当前配置 {os_label}", show_header=False, box=None)
        table.add_column("配置项", style="cyan")
        table.add_column("值", style="white")
        items = [
            ("表格路径 (excel_path)", cfg["excel_path"] or "(未设置)"),
            ("轮询间隔 (poll_seconds)", f"{cfg['poll_seconds']} 秒"),
            ("模拟运行 (dry_run)", "是 ✅" if cfg["dry_run"] else "否"),
            ("发送间隔 (send_interval)", f"{cfg['send_interval']} 秒"),
            ("每分钟最大条数 (max_per_minute)", str(cfg["max_per_minute"])),
        ]
        for k, v in items:
            table.add_row(k, v)
        console.print(table)
    else:
        for k, v in cfg.items():
            print(f"{k}: {v}")


def cmd_config_set(args):
    cfg = get_cfg()
    key_map = {
        "excel_path": ("表格路径", str),
        "poll_seconds": ("轮询间隔", int),
        "dry_run": ("模拟运行", lambda v: v.lower() in ("true", "1", "yes")),
        "send_interval": ("发送间隔", float),
        "max_per_minute": ("每分钟最大条数", int),
    }
    if args.key not in key_map:
        console.print(f"[red]未知配置项: {args.key}[/red]" if HAS_RICH else f"未知配置项: {args.key}")
        return
    label, type_fn = key_map[args.key]
    try:
        value = type_fn(args.value)
    except ValueError as e:
        console.print(f"[red]格式错误: {e}[/red]" if HAS_RICH else f"格式错误: {e}")
        return
    cfg[args.key] = value
    save_cfg(cfg)
    if HAS_RICH:
        console.print(f"[green]✅ 已更新 {label} = {value}[/green]")
    else:
        print(f"✅ 已更新 {label} = {value}")


def cmd_setup(_):
    if HAS_RICH:
        console.print(Panel("🛠️  首次配置向导", expand=False))
    cfg = get_cfg()

    if HAS_RICH:
        excel_path = Prompt.ask("[cyan]表格路径[/cyan]", default=cfg.get("excel_path", ""))
        send_interval = Prompt.ask("[cyan]发送间隔（秒）[/cyan]", default=str(cfg.get("send_interval", 5)))
        max_per_minute = Prompt.ask("[cyan]每分钟最大发送条数[/cyan]", default=str(cfg.get("max_per_minute", 8)))
        dry_run = Confirm.ask("[cyan]模拟运行[/cyan]", default=cfg.get("dry_run", False))
    else:
        excel_path = input(f"表格路径 [{cfg.get('excel_path', '')}]: ").strip() or cfg.get("excel_path", "")
        send_interval = input(f"发送间隔秒 [{cfg.get('send_interval', 5)}]: ").strip() or str(cfg.get("send_interval", 5))
        max_per_minute = input(f"每分钟最大条数 [{cfg.get('max_per_minute', 8)}]: ").strip() or str(cfg.get("max_per_minute", 8))
        dry_run = False

    cfg["excel_path"] = excel_path
    cfg["send_interval"] = float(send_interval)
    cfg["max_per_minute"] = int(max_per_minute)
    cfg["dry_run"] = dry_run
    save_cfg(cfg)
    print("\n✅ 配置已保存！")


def cmd_status(args):
    cfg = get_cfg()
    xlsx_path = Path(cfg.get("excel_path", "")).expanduser()
    if not xlsx_path.exists():
        if getattr(args, 'json', False):
            print(json.dumps({"error": f"表格不存在: {xlsx_path}"}))
        else:
            console.print(f"[red]表格不存在: {xlsx_path}[/red]" if HAS_RICH else f"表格不存在: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)

    waiting = sum(1 for t in tasks if t.status in ("", STATUS_WAITING) or not t.status)
    running = sum(1 for t in tasks if t.status == STATUS_RUNNING)
    success = sum(1 for t in tasks if t.status.startswith(STATUS_SUCCESS))
    failed  = sum(1 for t in tasks if t.status.startswith(STATUS_FAILED))

    if getattr(args, 'json', False):
        pending_tasks = [t for t in tasks if not t.status.startswith(STATUS_SUCCESS)]
        result = {
            "waiting": waiting,
            "running": running,
            "success": success,
            "failed": failed,
            "total": len(tasks),
            "excel_path": str(xlsx_path),
            "tasks": [
                {
                    "target": t.target,
                    "msg_type": t.msg_type,
                    "preview": (t.text or t.image_path or "")[:40],
                    "status": t.status or STATUS_WAITING,
                }
                for t in pending_tasks[:30]
            ],
        }
        print(json.dumps(result, ensure_ascii=False))
        return

    if HAS_RICH:
        table = Table(title=f"📊 任务概览 — {xlsx_path.name}", box=None)
        table.add_column("状态", style="white")
        table.add_column("数量", justify="right")
        table.add_row("⏳ 待发送", f"[yellow]{waiting}[/yellow]")
        table.add_row("🔄 发送中", f"[blue]{running}[/blue]")
        table.add_row("✅ 发送成功", f"[green]{success}[/green]")
        table.add_row("❌ 发送失败", f"[red]{failed}[/red]")
        table.add_row("📄 总计", str(len(tasks)))
        console.print(table)
    else:
        print(f"待发送: {waiting}, 发送中: {running}, 成功: {success}, 失败: {failed}, 总计: {len(tasks)}")


def _parse_tasks_from_json(json_str: str) -> list[Task]:
    """把 GUI 传来的 JSON 列表转成 Task 对象列表。"""
    import json as _json
    data = _json.loads(json_str)
    tasks = []
    for idx, item in enumerate(data):
        send_time = None
        if item.get("send_time"):
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    send_time = datetime.strptime(item["send_time"], fmt)
                    break
                except ValueError:
                    pass
        tasks.append(Task(
            row=0,  # 不关联 Excel 行
            app=str(item.get("app", "微信")).strip(),
            target=str(item.get("target", "")).strip(),
            msg_type=str(item.get("msg_type", "文字")).strip(),
            text=str(item.get("text", "")).strip(),
            image_path=str(item.get("image_path", "")).strip(),
            send_time=send_time,
            repeat=str(item.get("repeat", "")).strip(),
            status="",
        ))
    return tasks


def cmd_send(args):
    if not IS_MAC and not IS_WINDOWS:
        console.print("[red]❌ 当前系统不支持微信自动化[/red]" if HAS_RICH else "❌ 当前系统不支持微信自动化")
        return

    cfg = get_cfg()
    dry_run = cfg.get("dry_run", False)
    send_interval = cfg.get("send_interval", 5)
    max_per_minute = cfg.get("max_per_minute", 8)

    # ── GUI 传入的 --tasks-json 模式 ───────────────────────
    if getattr(args, "tasks_json", None):
        tasks = _parse_tasks_from_json(args.tasks_json)
        pending = [t for t in tasks if should_send(t, datetime.now())]
        if not pending:
            print("没有需要发送的任务")
            return
        platform_note = "🍎 macOS" if IS_MAC else "🪟 Windows"
        print(f"开始发送 {len(pending)} 条任务 [{platform_note}]...")
        if dry_run:
            print("⚠️  模拟运行模式")
        sent_times: list[datetime] = []
        success_count, fail_count = 0, 0
        for i, task in enumerate(pending):
            window_start = datetime.now() - timedelta(minutes=1)
            sent_times = [t for t in sent_times if t > window_start]
            if len(sent_times) >= max_per_minute:
                sleep_secs = 60 - (datetime.now() - sent_times[0]).total_seconds()
                if sleep_secs > 0:
                    time.sleep(sleep_secs)
                    sent_times = [t for t in sent_times if t > datetime.now() - timedelta(minutes=1)]
            status_str = ''
            try:
                if not dry_run:
                    validate_task(task)
                    call_sender(task.target, task.msg_type, task.text, task.image_path)
                sent_times.append(datetime.now())
                success_count += 1
                status_str = f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}] ✅"
            except Exception as e:
                fail_count += 1
                status_str = f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}] ❌ {e}"
            print(status_str)
            if i < len(pending) - 1 and not dry_run:
                time.sleep(send_interval)
        print(f"\n完成！成功 {success_count} 条，失败 {fail_count} 条")
        if fail_count > 0 or success_count == 0:
            sys.exit(1)
        return

    # ── 原有的 Excel 模式 ───────────────────────────────────
    xlsx_path = Path(cfg.get("excel_path", "")).expanduser()
    if not xlsx_path.exists():
        console.print(f"[red]表格不存在: {xlsx_path}[/red]" if HAS_RICH else f"表格不存在: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)
    now = datetime.now()

    pending = [t for t in tasks if should_send(t, now)]
    if not pending:
        print("没有需要发送的任务")
        return

    platform_note = "🍎 macOS" if IS_MAC else "🪟 Windows"
    print(f"开始发送 {len(pending)} 条任务 [{platform_note}]...")
    if dry_run:
        print("⚠️  模拟运行模式，不会真实发送")

    sent_times: list[datetime] = []
    success_count, fail_count = 0, 0

    for i, task in enumerate(pending):
        window_start = now - timedelta(minutes=1)
        sent_times = [t for t in sent_times if t > window_start]
        if len(sent_times) >= max_per_minute:
            sleep_secs = 60 - (datetime.now() - sent_times[0]).total_seconds()
            if sleep_secs > 0:
                print(f"达到每分钟 {max_per_minute} 条限制，等待 {sleep_secs:.0f}s...")
                time.sleep(sleep_secs)
                sent_times = [t for t in sent_times if t > datetime.now() - timedelta(minutes=1)]

        set_status(ws, cols, task.row, STATUS_RUNNING)
        wb.save(xlsx_path)

        status_str = ''
        try:
            if not dry_run:
                validate_task(task)
                call_sender(task.target, task.msg_type, task.text, task.image_path)
            status = f"{STATUS_SUCCESS} {now.strftime('%H:%M:%S')}"
            set_status(ws, cols, task.row, status)
            sent_times.append(datetime.now())
            success_count += 1
            status_str = f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}] ✅"
        except Exception as e:
            status = f"{STATUS_FAILED}: {e}"
            set_status(ws, cols, task.row, status)
            fail_count += 1
            status_str = f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}] ❌ {e}"
        print(status_str)

        wb.save(xlsx_path)

        if i < len(pending) - 1 and not dry_run:
            time.sleep(send_interval)

    print(f"\n完成！成功 {success_count} 条，失败 {fail_count} 条")
    if fail_count > 0 or (success_count == 0 and pending):
        sys.exit(1)


def _load_gui_tasks() -> list[dict]:
    """从 gui_tasks.json 加载 GUI 任务列表（不依赖 openpyxl）。"""
    f = CFG_PATH.parent / "gui_tasks.json"
    if not f.exists():
        return []
    try:
        with open(f, "r", encoding="utf-8") as fh:
            return json.load(fh) or []
    except Exception:
        return []


def _save_gui_tasks(tasks: list[dict]):
    """保存任务列表到 gui_tasks.json。"""
    f = CFG_PATH.parent / "gui_tasks.json"
    CFG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(f, "w", encoding="utf-8") as fh:
        json.dump(tasks, fh, ensure_ascii=False, indent=2)


def cmd_daemon(args):
    if not IS_MAC and not IS_WINDOWS:
        print("❌ 当前系统不支持微信自动化")
        return

    cfg = get_cfg()
    poll_seconds = int(cfg.get("poll_seconds", 15))
    dry_run = cfg.get("dry_run", False)
    send_interval = cfg.get("send_interval", 5)
    max_per_minute = cfg.get("max_per_minute", 8)

    print(f"守护进程模式启动 | 轮询间隔 {poll_seconds}s | dry_run={dry_run}")
    sent_times: list[datetime] = []

    while True:
        try:
            tasks = _load_gui_tasks()
            now = datetime.now()
            changed = False

            for task_dict in tasks:
                status = task_dict.get("status", "") or ""
                # 跳过已成功的非重复任务
                if status.startswith(STATUS_SUCCESS) and not task_dict.get("repeat"):
                    continue

                # 解析 send_time
                send_time = None
                raw_st = task_dict.get("send_time", "") or ""
                if raw_st:
                    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                        try:
                            send_time = datetime.strptime(raw_st, fmt)
                            break
                        except ValueError:
                            pass

                # 没有 send_time 的任务留着手动处理（守护进程跳过）
                if send_time is None:
                    continue

                if now < send_time:
                    if not status:
                        task_dict["status"] = STATUS_WAITING
                        changed = True
                    continue

                # 频率控制
                window_start = now - timedelta(minutes=1)
                sent_times = [t for t in sent_times if t > window_start]
                if len(sent_times) >= max_per_minute:
                    time.sleep(1)
                    continue

                task_dict["status"] = STATUS_RUNNING
                changed = True

                try:
                    if not dry_run:
                        validate_task(Task(
                            row=0, app=task_dict.get("app", "微信"), target=task_dict.get("target", ""),
                            msg_type=task_dict.get("msg_type", "文字"), text=task_dict.get("text", ""),
                            image_path=task_dict.get("image_path", ""),
                            send_time=send_time, repeat=task_dict.get("repeat", ""), status=STATUS_RUNNING,
                        ))
                        call_sender(
                            task_dict.get("target", ""),
                            task_dict.get("msg_type", "文字"),
                            task_dict.get("text", ""),
                            task_dict.get("image_path", ""),
                        )
                    task_dict["status"] = f"{STATUS_SUCCESS} {now.strftime('%m-%d %H:%M')}"
                    sent_times.append(datetime.now())
                    print(f"✅ → {task_dict.get('target')} [{task_dict.get('msg_type')}]", flush=True)
                    time.sleep(send_interval)
                except Exception as e:
                    task_dict["status"] = f"{STATUS_FAILED}: {e}"
                    print(f"❌ → {task_dict.get('target')}: {e}", flush=True)

            if changed:
                _save_gui_tasks(tasks)

        except Exception as e:
            print(f"错误: {e}", flush=True)

        time.sleep(poll_seconds)


def cmd_template(_):
    print("Excel 模板列名：")
    cols = [
        ("* 应用", "固定填 微信"),
        ("* 联系人/群聊", "对方微信名或群名"),
        ("* 消息类型", "文字 / 图片 / 文字+图片"),
        ("* 文字内容", "要发送的文本"),
        ("图片路径", "本地图片绝对路径"),
        ("发送时间", "格式 2025-01-01 14:30"),
        ("重复", "daily / weekly / workday / 空"),
        ("状态", "自动填充，勿手动编辑"),
    ]
    for k, v in cols:
        print(f"  {k}: {v}")


def main():
    parser = argparse.ArgumentParser(description="微信批量发送助手 CLI (macOS / Windows)")
    sub = parser.add_subparsers(dest="cmd")

    p_config = sub.add_parser("config", help="查看 / 修改配置")
    p_config.add_argument("key", nargs="?", help="配置项名称")
    p_config.add_argument("value", nargs="?", help="新的配置值")
    p_config.add_argument("--json", action="store_true", help="JSON 输出（供 GUI 调用）")
    p_config.set_defaults(func=lambda a: cmd_config_set(a) if a.key and a.value else cmd_config_show(a))

    p_status = sub.add_parser("status", help="查看任务状态")
    p_status.add_argument("--json", action="store_true", help="JSON 输出")
    p_status.set_defaults(func=cmd_status)

    sub.add_parser("setup", help="交互式配置向导").set_defaults(func=cmd_setup)
    p_send = sub.add_parser("send", help="立即发送")
    p_send.add_argument("--tasks-json", help="JSON 格式任务数据（绕过 Excel 文件）")
    p_send.set_defaults(func=cmd_send)
    sub.add_parser("daemon", help="守护进程模式").set_defaults(func=cmd_daemon)
    sub.add_parser("template", help="查看表格模板").set_defaults(func=cmd_template)

    args = parser.parse_args()

    if args.cmd is None:
        parser.print_help()
        return

    args.func(args)


if __name__ == "__main__":
    main()
