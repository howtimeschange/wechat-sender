#!/usr/bin/env python3
"""
微信批量发送助手 — Windows 版（pywinauto 方案）
依赖: pip install pywinauto psutil pyperclip Pillow openpyxl PyYAML rich
核心优势：keyboard.send_keys() 用 SendMessage WM_SETTEXT 直接写文本到控件，
         不走剪贴板，不触发 WeChat 剪贴板监控，无防抖问题。
用法: python scripts/wechat_send_win.py [--dry]
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional


def _real_home() -> Path:
    """返回真实用户 home 目录，绕过 Python pkg 改写的 $HOME（macOS）"""
    if sys.platform == "win32":
        return Path(os.environ.get("USERPROFILE",
                                  os.environ.get("HOMEDRIVE", "C:\\") + "\\Users\\" + os.environ.get("USERNAME", "")))
    import pwd
    return Path(pwd.getpwuid(os.getuid()).pw_dir)


# ─── 依赖安装 ──────────────────────────────────────────
def _install_and_import(module, package=None):
    """尝试导入，失败则自动安装"""
    try:
        return __import__(module)
    except ImportError:
        package = package or module
        pkg = {"pywinauto": "pywinauto", "psutil": "psutil",
               "pyperclip": "pyperclip", "Pillow": "Pillow",
               "openpyxl": "openpyxl", "PyYAML": "PyYAML", "rich": "rich"}
        pypi_name = pkg.get(package, package)
        print(f"[信息] 正在安装 {pypi_name}...")
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", pypi_name, "--quiet"],
            capture_output=True, text=True
        )
        if r.returncode != 0:
            print(f"[ERROR] pip install {pypi_name} 失败: {r.stderr}", file=sys.stderr)
            sys.exit(1)
        return __import__(module)


# 提前导入（自动安装）
pywinauto_mod = _install_and_import("pywinauto")
from pywinauto import Application, keyboard, mouse, Desktop
from pywinauto.timings import wait_until_passes
import psutil
_pyperclip = _install_and_import("pyperclip")
_pillow = _install_and_import("PIL", "Pillow")
Image = _pillow.Image
_pyperclip_extras = {"pywin32": None, "win32clipboard": None, "win32con": None}
try:
    import win32clipboard
    import win32con
    _pyperclip_extras["pywin32"] = True
except ImportError:
    pass


# ─── UTF-8 编码保障 ────────────────────────────────────
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# ─── OS 版本检测（Win10 vs Win11 延迟系数）────────────────
def _get_os_timing() -> dict:
    """
    检测 Windows 版本，返回各阶段的延迟配置。
    Win11 build >= 22000，消息循环更快，用较短延迟。
    Win10 build < 22000，自绘控件 + 消息队列积压，需要保守延迟。
    """
    try:
        import platform as _plat
        build = int(_plat.version().split(".")[-1])
    except Exception:
        build = 0

    if build >= 22000:
        # Win11：标准 UIA 控件响应快
        return {
            "is_win11": True,
            "focus_delay": 0.25,        # set_focus 后等待
            "search_key_delay": 0.25,   # ^f / ^k 后等待
            "search_result_timeout": 2.0,
            "chat_ready_timeout": 2.0,
            "esc_delay": 0.1,
            "input_retry_base": 0.1,
        }
    else:
        # Win10：自绘控件，消息队列慢，全部×1.6
        return {
            "is_win11": False,
            "focus_delay": 0.4,
            "search_key_delay": 0.4,
            "search_result_timeout": 4.0,
            "chat_ready_timeout": 3.5,
            "esc_delay": 0.2,
            "input_retry_base": 0.15,
        }

_OS_TIMING = _get_os_timing()


# ─── 剪贴板工具 ────────────────────────────────────────
def _clipboard_copy(text: str):
    """写入文本到剪贴板（仅在需要粘贴图片时使用）"""
    if not text:
        return
    try:
        _pyperclip.copy(text)
    except Exception as e:
        print(f"[WARN] 剪贴板写入失败: {e}", file=sys.stderr)


def _clipboard_clear():
    try:
        _pyperclip.copy("")
    except Exception:
        pass


# ─── pywinauto 核心工具 ─────────────────────────────────

BACKEND = "uia"   # 默认 UIA；win32 作为兜底


def _log(msg: str):
    if VERBOSE:
        print(msg)


def _window_area(rect) -> int:
    try:
        return max(0, rect.width() * rect.height())
    except Exception:
        return 0


def _safe_enum_windows(backend: str, timeout: float = 2.0):
    """线程安全枚举桌面窗口，避免卡死"""
    result = {"windows": None}

    def _worker():
        try:
            result["windows"] = Desktop(backend=backend).windows()
        except Exception:
            result["windows"] = []

    import threading
    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout)
    if t.is_alive():
        _log(f"[WARN] 枚举窗口超时（backend={backend}）")
        return []
    return result["windows"] or []


def _find_wechat_window():
    """枚举桌面顶层窗口，找到最可能是微信主窗口的那个"""
    top_windows = _safe_enum_windows(BACKEND, timeout=2.0)
    if not top_windows:
        alt = "win32" if BACKEND == "uia" else "uia"
        top_windows = _safe_enum_windows(alt, timeout=2.0)
    if not top_windows:
        return None

    candidates = []
    for w in top_windows:
        try:
            ei = w.element_info
            name = (ei.name or "")
            class_name = (ei.class_name or "")
            pid = getattr(ei, "process_id", None)
            proc_name = ""
            if pid:
                try:
                    proc_name = (psutil.Process(pid).name() or "").lower()
                except Exception:
                    pass
            # 过滤非微信进程
            if not any(x in name for x in ("微信", "WeChat", "Weixin")):
                continue
            if proc_name not in ("weixin.exe", "wechat.exe", "wechat.exe", ""):
                if proc_name:  # 有进程名但不是微信
                    continue
            # 评分
            score = 0
            if class_name in ("WeChatMainWndForPC", "WeChatMainWndForPC64",
                              "WeChatMainWnd", "MainWindow", "Window"):
                score += 5
            area = _window_area(ei.rectangle)
            score += min(5, area // (800 * 600))
            candidates.append((score, area, w))
        except Exception:
            continue

    if not candidates:
        return None
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    chosen = candidates[0][2]
    try:
        ei = chosen.element_info
        _log(f"[INFO] 选择窗口：title='{ei.name}' class='{ei.class_name}' "
             f"area={_window_area(ei.rectangle)} pid={getattr(ei, 'process_id', '')}")
    except Exception:
        pass
    return chosen


def ensure_wechat_running(start_if_needed: bool = True, timeout: float = 20.0):
    """确保微信已运行，无窗口则尝试启动"""
    win = _find_wechat_window()
    if win is not None:
        _log("[INFO] 已找到微信主窗口")
        return

    if not start_if_needed:
        raise RuntimeError("微信未运行，请先启动微信")

    # 尝试启动
    candidates = [
        os.path.expandvars(r"%LOCALAPPDATA%\Tencent\WeChat\WeChat.exe"),
        os.path.expandvars(r"%PROGRAMFILES%\Tencent\WeChat\WeChat.exe"),
        os.path.expandvars(r"%PROGRAMFILES(X86)%\Tencent\WeChat\WeChat.exe"),
        "WeChat.exe",
    ]
    started = False
    for exe in candidates:
        if not os.path.isfile(exe) if os.path.isabs(exe) else True:
            # 只检查绝对路径
            if os.path.isabs(exe) and not os.path.isfile(exe):
                continue
        try:
            _log(f"[INFO] 启动: {exe}")
            subprocess.Popen([exe], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            started = True
            break
        except FileNotFoundError:
            continue

    if not started:
        # 最后一个候选：直接用 start 命令
        try:
            subprocess.Popen(["cmd", "/c", "start", "", "WeChat.exe"],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            started = True
        except Exception:
            pass

    # 等待窗口出现
    def _connected():
        return _find_wechat_window() is not None

    try:
        wait_until_passes(timeout, 1.0, _connected)
    except Exception:
        raise RuntimeError(f"等待微信窗口超时（{timeout}s），请手动启动微信")


def attach_wechat(timeout: float = 20.0):
    """附着到微信窗口，返回 (app, main_window_wrapper)。
    返回的 main_win 是真正的 wrapper 对象（不是懒加载 spec），
    避免跨调用时 element_info 失效导致 '__dict__' 错误。
    """
    ensure_wechat_running(start_if_needed=True, timeout=timeout)
    chosen = _find_wechat_window()
    if chosen is None:
        raise RuntimeError("无法找到微信主窗口")

    # 先尝试 UIA backend，失败则降级 win32
    for backend in (BACKEND, "win32"):
        try:
            app = Application(backend=backend)
            try:
                app.connect(handle=chosen.handle, timeout=3)
            except Exception:
                try:
                    app.connect(title_re="微信|WeChat|Weixin", timeout=3)
                except Exception:
                    app.connect(path_re=r"Weixin\.exe|WeChat\.exe", timeout=3)

            # wrapper_object() 可能返回 None（UIA 节点不可用时）
            spec = app.top_window()
            wrapper = None
            try:
                wrapper = spec.wrapper_object()
            except Exception:
                pass

            if wrapper is None:
                # 尝试通过标题直接定位
                try:
                    wrapper = app.window(title_re="微信|WeChat|Weixin").wrapper_object()
                except Exception:
                    pass

            if wrapper is None:
                _log(f"[WARN] backend={backend} wrapper_object() 返回 None，尝试下一个")
                continue

            # 验证确实是微信
            try:
                name = (wrapper.element_info.name or "").lower()
                if not ("微信" in name or "wechat" in name or "weixin" in name):
                    alt = app.window(title_re="微信|WeChat|Weixin").wrapper_object()
                    if alt is not None:
                        wrapper = alt
            except Exception:
                pass

            # 还原最小化
            try:
                if getattr(wrapper, "is_minimized", None) and wrapper.is_minimized():
                    _log("[INFO] 还原最小化的微信窗口")
                    wrapper.restore()
            except Exception:
                pass

            # 聚焦
            _log(f"[INFO] 聚焦微信窗口（backend={backend}）")
            try:
                wrapper.set_focus()
            except Exception:
                pass

            time.sleep(_OS_TIMING["focus_delay"])
            return app, wrapper

        except Exception as e:
            _log(f"[WARN] backend={backend} 附着失败: {e}")
            continue

    raise RuntimeError("所有 backend 均无法附着微信窗口，请确认微信已登录并在前台运行")


# ─── 搜索 & 发送 ────────────────────────────────────────

def _focus_search_edit(main_win):
    """尝试直接聚焦搜索框 Edit 控件"""
    try:
        edits = main_win.descendants(control_type="Edit")
    except Exception:
        edits = []
    for edit in edits[:5]:
        name = (getattr(edit.element_info, "name", "") or "").lower()
        if ("search" in name or "搜索" in name or "查找" in name or name == ""):
            try:
                edit.set_focus()
                _log(f"[INFO] 已聚焦搜索框 Edit，name='{name}'")
                return True
            except Exception:
                continue
    return False


def _get_win_rect(win):
    """兼容 UIA/win32 两种 backend 获取窗口矩形。"""
    try:
        # win32 backend: wrapper 有 .rectangle() 方法
        r = win.rectangle()
        return r
    except Exception:
        pass
    try:
        # UIA backend: element_info.rectangle 属性
        return win.element_info.rectangle
    except Exception:
        return None


def _click_search_area(main_win):
    """点击微信左侧搜索框区域（坐标法，Win10/Win11 通用）。
    先尝试 UIA 控件聚焦，失败则降级到坐标点击。
    """
    if _focus_search_edit(main_win):
        return True

    try:
        rect = _get_win_rect(main_win)
        if rect is None:
            return False
        cx = rect.left + 175
        cy = rect.top + 48
        _log(f"[INFO] 坐标点击搜索框区域 ({cx}, {cy})")
        mouse.click(button="left", coords=(cx, cy))
        time.sleep(0.15)
        return True
    except Exception as e:
        _log(f"[WARN] 坐标点击失败: {e}")
        return False


def focus_search_and_open_chat(main_win, friend_name: str):
    """聚焦全局搜索框，输入好友名，回车打开聊天。
    OS 感知版：Win10/Win11 使用不同延迟参数。
    """
    t = _OS_TIMING
    try:
        main_win.set_focus()
    except Exception:
        pass
    time.sleep(t["focus_delay"])

    # ── 退出当前聊天状态，回到主界面 ──
    try:
        rect = _get_win_rect(main_win)
        if rect is not None:
            lx = rect.left + 80
            ly = rect.top + int((rect.bottom - rect.top) / 3)
            _log(f"[INFO] 点击会话列表区退出聊天 ({lx}, {ly})")
            mouse.click(button="left", coords=(lx, ly))
            time.sleep(t["esc_delay"])
        else:
            raise ValueError("无法获取窗口矩形")
    except Exception:
        keyboard.send_keys("{ESC}")
        time.sleep(t["esc_delay"])
        keyboard.send_keys("{ESC}")
        time.sleep(t["esc_delay"])

    # ── 聚焦搜索框 ──
    _click_search_area(main_win)
    time.sleep(t["search_key_delay"])

    if not _focus_search_edit(main_win):
        keyboard.send_keys("^f")
        time.sleep(t["search_key_delay"])

    keyboard.send_keys("^a{BACKSPACE}")
    time.sleep(0.15)

    _log(f"[INFO] 输入搜索词：{friend_name}")
    keyboard.send_keys(friend_name, with_spaces=True)

    _wait_for_search_result(main_win, friend_name, timeout=t["search_result_timeout"])

    _log("[INFO] 回车打开聊天")
    keyboard.send_keys("{ENTER}")

    _wait_for_chat_ready(main_win, timeout=t["chat_ready_timeout"])


def _wait_for_search_result(main_win, friend_name: str, timeout: float = 3.0):
    """自适应等待搜索结果出现。
    策略：轮询 ListItem / DataItem 类控件，出现且数量稳定后认为列表加载完毕。
    兜底：超时后用固定 0.5s 延迟。
    """
    _log(f"[INFO] 等待搜索结果（最多 {timeout}s）")
    deadline = time.time() + timeout
    prev_count = -1
    stable_ticks = 0

    while time.time() < deadline:
        try:
            # 检查搜索结果列表（ListItem / DataItem 通常是搜索结果条目）
            items = main_win.descendants(control_type="ListItem")
            if not items:
                items = main_win.descendants(control_type="DataItem")
            count = len(items)
            if count > 0:
                if count == prev_count:
                    stable_ticks += 1
                    if stable_ticks >= 2:   # 连续 2 次（≈160ms）结果稳定
                        _log(f"[INFO] 搜索结果稳定，共 {count} 项")
                        time.sleep(0.1)
                        return
                else:
                    stable_ticks = 0
                prev_count = count
        except Exception:
            pass
        time.sleep(0.08)

    _log("[WARN] 搜索结果等待超时，用固定延迟兜底")
    time.sleep(0.5)


def _wait_for_chat_ready(main_win, timeout: float = 2.5):
    """自适应等待聊天输入框就绪。
    策略：找所有 Edit/Document/RichEdit 控件，取面积最大的那个。
    面积最大的才是聊天输入框（搜索框面积小得多）。
    Win10 上 threshold 降低到 5000 避免漏判。
    """
    _log(f"[INFO] 等待聊天窗口就绪（最多 {timeout}s）")
    # Win10 聊天输入框面积阈值更小（窗口可能不是最大化）
    area_threshold = 5000 if not _OS_TIMING["is_win11"] else 10000

    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            ctrls = main_win.descendants()
            candidates = []
            for c in ctrls:
                try:
                    ei = c.element_info
                    ct = getattr(ei, "control_type", "") or ""
                    cn = getattr(ei, "class_name", "") or ""
                    if ct not in ("Edit", "Document") and "RichEdit" not in cn:
                        continue
                    rect = getattr(ei, "rectangle", None)
                    if rect:
                        area = _window_area(rect)
                        if area > area_threshold:
                            candidates.append((area, c))
                except Exception:
                    continue

            if candidates:
                # 取面积最大的，排除搜索框（搜索框通常 < 20000）
                candidates.sort(key=lambda x: x[0], reverse=True)
                largest_area = candidates[0][0]
                _log(f"[INFO] 最大输入控件面积={largest_area}，等待阈值={area_threshold}")
                if largest_area > area_threshold * 3:  # 聊天框 >> 搜索框
                    _log("[INFO] 聊天输入框已就绪")
                    time.sleep(0.15)
                    return
        except Exception:
            pass
        time.sleep(0.1)

    _log("[WARN] 等待超时，使用固定延迟兜底")
    time.sleep(0.5 if not _OS_TIMING["is_win11"] else 0.3)


def _focus_message_input(main_win):
    """尝试聚焦聊天输入框，返回是否成功"""
    try:
        ctrls = main_win.descendants()
    except Exception:
        ctrls = []

    try:
        rect_win = _get_win_rect(main_win)
        bottom_win = rect_win.bottom if rect_win else 99999
    except Exception:
        bottom_win = 99999

    scored = []
    for c in ctrls:
        try:
            ei = c.element_info
            ct = getattr(ei, "control_type", "") or ""
            cn = getattr(ei, "class_name", "") or ""
            nm = getattr(ei, "name", "") or ""
            rect = getattr(ei, "rectangle", None)
            if ct not in ("Edit", "Document", "Text") and "RichEdit" not in cn:
                continue
            if rect is None:
                continue
            area = _window_area(rect)
            distance = max(0, bottom_win - rect.bottom)
            score = area - distance * 10
            scored.append((score, rect, c, ct, cn, nm))
        except Exception:
            continue

    if not scored:
        _log("[WARN] 未找到候选输入控件")
        return False

    scored.sort(key=lambda x: x[0], reverse=True)
    for score, rect, ctrl, ct, cn, nm in scored[:5]:
        try:
            _log(f"[INFO] 尝试聚焦输入控件 type={ct} class={cn} name={nm}")
            ctrl.set_focus()
            return True
        except Exception:
            # set_focus 失败，尝试点击控件中心
            x = int((rect.left + rect.right) / 2)
            y = int((rect.top + rect.bottom) / 2)
            mouse.click(button="left", coords=(x, y))
            return True
    return False


def _click_bottom_chat_area(main_win, clicks: int = 3):
    """点击聊天窗口底部区域以获取焦点"""
    try:
        rect = _get_win_rect(main_win)
        if rect is None:
            return
        cx = int((rect.left + rect.right) / 2)
        for i in range(clicks):
            y = int(rect.bottom - 80 - i * 40)
            _log(f"[DEBUG] 点击聊天底部区域 ({cx}, {y})")
            mouse.click(button="left", coords=(cx, y))
            time.sleep(0.1)
    except Exception as e:
        _log(f"[WARN] 点击聊天底部失败: {e}")


def send_message_to_current_chat(main_win, message: str,
                                 delay: float = 0.15,
                                 press_enter_to_send: bool = True,
                                 use_paste: bool = False):
    """
    向当前聊天窗口输入消息并发送。
    use_paste=True 时用 Ctrl+V（走剪贴板），False 时直接打字。
    OS 感知版：Win10 重试间隔更长，Win11 更快。
    """
    t = _OS_TIMING
    retry_base = t["input_retry_base"]

    # 聚焦输入框（最多重试 5 次，兼容 Win10 慢机型）
    focused = False
    for attempt in range(5):
        if _focus_message_input(main_win):
            focused = True
            break
        _click_bottom_chat_area(main_win, clicks=2)
        time.sleep(retry_base + attempt * retry_base)  # 逐步加大等待，Win10 更保守
    if not focused:
        _log("[WARN] 未能聚焦输入框，继续尝试发送")

    time.sleep(0.08)
    keyboard.send_keys("{END}")   # 确保光标在末尾
    time.sleep(0.1)

    # 输入消息：优先直接打字（不走剪贴板，无防抖问题）
    _log(f"[INFO] 输入消息：{message[:20]}{'...' if len(message) > 20 else ''}")
    if use_paste:
        _clipboard_copy(message)
        time.sleep(0.08)
        keyboard.send_keys("^v")
    else:
        keyboard.send_keys(message, with_spaces=True)

    time.sleep(delay)

    # 发送
    if press_enter_to_send:
        keyboard.send_keys("{ENTER}")
    else:
        keyboard.send_keys("^{ENTER}")
    time.sleep(delay)


def search_contact(name: str, max_retries: int = 2):
    """搜索并打开与指定联系人的聊天窗口（仅供独立测试使用；批量发送请用 call_send）"""
    app, main_win = attach_wechat()
    for attempt in range(max_retries):
        _log(f"[INFO] search_contact attempt {attempt + 1}: {name}")
        focus_search_and_open_chat(main_win, name)
        return
    raise RuntimeError(f"未找到联系人 [{name}]，请手动确认微信窗口状态")


def send_text(text: str, main_win=None):
    """发送文字消息（直接打字，不走剪贴板）"""
    send_message_to_current_chat(main_win, text, delay=0.12,
                                 press_enter_to_send=True, use_paste=False)


def send_image(image_path: str):
    """发送图片（必须走剪贴板）"""
    img = Image.open(image_path).convert("RGB")
    _clipboard_clear()
    try:
        if "pywin32" in _pyperclip_extras and win32clipboard:
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_BITMAP, img)
            win32clipboard.CloseClipboard()
        else:
            # 兜底：用 pyperclip 写文件路径
            _clipboard_copy(image_path)
    except Exception as e:
        print(f"[WARN] 图片剪贴板写入失败: {e}", file=sys.stderr)
        _clipboard_copy(image_path)

    time.sleep(0.15)
    keyboard.send_keys("^v")
    time.sleep(0.25)
    keyboard.send_keys("{ENTER}")


def send_text_with_image(text: str, image_path: str):
    """发送文字+图片"""
    _clipboard_copy(text)
    time.sleep(0.05)
    keyboard.send_keys("^v")
    time.sleep(0.2)
    send_image(image_path)


def call_send(target: str, msg_type: str, text: str, image_path: str):
    """统一发送入口（由 cli.py 调用）。
    进程内批量调用时复用同一个 app/main_win wrapper，不重新附着微信。
    """
    global _cached_app, _cached_main_win

    # 懒加载：首次调用时附着微信，之后复用
    if _cached_main_win is None:
        _cached_app, _cached_main_win = attach_wechat()
    else:
        # 验证 wrapper 仍然有效（通过 handle 判断，不触发 element_info 懒求值）
        try:
            handle = _cached_main_win.handle
            if not handle:
                raise ValueError("handle 无效")
        except Exception:
            _log("[INFO] 缓存窗口已失效，重新附着微信")
            _cached_app, _cached_main_win = attach_wechat()

    main_win = _cached_main_win
    focus_search_and_open_chat(main_win, target)

    if msg_type == "文字":
        send_message_to_current_chat(main_win, text, delay=0.15,
                                     press_enter_to_send=True, use_paste=False)
    elif msg_type == "图片":
        send_image(image_path)
    elif msg_type == "文字+图片":
        _clipboard_copy(text)
        time.sleep(0.08)
        keyboard.send_keys("^v")
        time.sleep(0.2)
        send_image(image_path)
    else:
        raise ValueError(f"不支持的消息类型: {msg_type}")


# ─── 配置（从外部 config.json 读取）──────────────────────

ROOT = Path(__file__).resolve().parents[1]
CFG_PATH = _real_home() / ".wechat-sender" / "config.json"
XLSX_PATH = None

SHEET_TASKS = "发送任务"
HEADER_ROW = 2

STATUS_WAITING = "待发送"
STATUS_RUNNING = "发送中"
STATUS_SUCCESS = "发送成功"
STATUS_FAILED = "发送失败"

COL = {
    "seq": "#",
    "app": "* 应用",
    "target": "* 联系人/群聊",
    "msg_type": "* 消息类型",
    "text": "* 文字内容",
    "image": "图片路径",
    "send_time": "发送时间",
    "repeat": "重复",
    "remark": "备注",
    "status": "状态",
}


def load_cfg() -> dict:
    if not CFG_PATH.exists():
        return {}
    with open(CFG_PATH, "r", encoding="utf-8") as f:
        return json.load(f) or {}


@dataclass
class Task:
    row: int
    app: str
    target: str
    msg_type: str
    text: str
    image_path: str
    send_time: Optional[datetime]
    repeat: str
    status: str


def find_columns(ws):
    cols = {}
    for c in range(1, ws.max_column + 1):
        title = (ws.cell(HEADER_ROW, c).value or "").strip()
        if title:
            cols[title] = c
    missing = [v for k, v in COL.items() if v not in cols and not k.startswith("seq")]
    if missing:
        raise RuntimeError(f"模板缺少列: {missing}")
    return cols


def read_tasks(ws, cols) -> list[Task]:
    tasks = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        app = str(ws.cell(r, cols[COL["app"]]).value or "").strip()
        target = str(ws.cell(r, cols[COL["target"]]).value or "").strip()
        msg_type = str(ws.cell(r, cols[COL["msg_type"]]).value or "").strip()
        text = str(ws.cell(r, cols[COL["text"]]).value or "").strip()
        image_path = str(ws.cell(r, cols[COL["image"]]).value or "").strip()
        send_time_raw = ws.cell(r, cols[COL["send_time"]]).value
        repeat_raw = ws.cell(r, cols[COL["repeat"]]).value
        status = str(ws.cell(r, cols[COL["status"]]).value or "").strip()
        if not any([app, target, msg_type, text, image_path, send_time_raw, repeat_raw, status]):
            continue
        send_time = send_time_raw if isinstance(send_time_raw, datetime) else None
        tasks.append(Task(
            row=r, app=app, target=target, msg_type=msg_type,
            text=text, image_path=image_path,
            send_time=send_time,
            repeat=str(repeat_raw).strip() if repeat_raw else "",
            status=status,
        ))
    return tasks


def set_status(ws, cols, row: int, status: str):
    ws.cell(row, cols[COL["status"]]).value = status


def should_send(task: Task, now: datetime) -> bool:
    if task.status.startswith(STATUS_SUCCESS) and not task.repeat:
        return False
    if task.send_time is None:
        return True
    return now >= task.send_time


# ─── 主发送逻辑 ─────────────────────────────────────────

def batch_send(dry_run: bool = False, send_interval: float = 5,
               max_per_minute: int = 8, xlsx_path: str = ""):
    import openpyxl

    if not xlsx_path:
        raise ValueError("未设置 excel_path，请先运行 python app/cli.py setup 配置表格路径")
    xlsx_path = Path(xlsx_path).expanduser()
    if not xlsx_path.exists():
        raise RuntimeError(f"表格不存在: {xlsx_path}")

    print(f"📋 读取表格: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)
    now = datetime.now()

    pending = [t for t in tasks if should_send(t, now)]
    if not pending:
        print("⏳ 没有需要发送的任务")
        return

    print(f"📤 开始发送 {len(pending)} 条任务（pywinauto 方案）...")
    if dry_run:
        print("⚠️  模拟运行模式，不会真实发送")

    sent_times: list[datetime] = []
    success_count, fail_count = 0, 0

    for i, task in enumerate(pending):
        # 频率控制
        window_start = now - timedelta(minutes=1)
        sent_times = [t for t in sent_times if t > window_start]
        if len(sent_times) >= max_per_minute:
            sleep_secs = 60 - (datetime.now() - sent_times[0]).total_seconds()
            if sleep_secs > 0:
                print(f"⏳ 达到每分钟 {max_per_minute} 条限制，等待 {sleep_secs:.0f}s...")
                time.sleep(sleep_secs)
                sent_times = [t for t in sent_times
                              if t > datetime.now() - timedelta(minutes=1)]

        print(f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}]", end="")

        set_status(ws, cols, task.row, STATUS_RUNNING)
        wb.save(xlsx_path)

        try:
            if not dry_run:
                if task.app and task.app != "微信":
                    raise ValueError(f"不支持的应用: {task.app}（当前仅支持微信）")
                if not task.target:
                    raise ValueError("联系人/群聊不能为空")
                if task.msg_type not in {"文字", "图片", "文字+图片"}:
                    raise ValueError(f"不支持的消息类型: {task.msg_type}")
                if task.msg_type in {"文字", "文字+图片"} and not task.text:
                    raise ValueError("文字内容不能为空")
                call_send(task.target, task.msg_type, task.text, task.image_path)

            status = f"{STATUS_SUCCESS} {datetime.now().strftime('%H:%M:%S')}"
            set_status(ws, cols, task.row, status)
            sent_times.append(datetime.now())
            success_count += 1
            print(" ✅")
        except Exception as e:
            set_status(ws, cols, task.row, f"{STATUS_FAILED}: {e}")
            fail_count += 1
            print(f" ❌ {e}")

        wb.save(xlsx_path)

        if i < len(pending) - 1 and not dry_run:
            time.sleep(send_interval)

    print(f"\n✅ 完成！成功 {success_count} 条，失败 {fail_count} 条")


# ─── 全局开关 ───────────────────────────────────────────
VERBOSE = False

# 模块级缓存：批量发送时复用同一个微信窗口引用，不重新附着
_cached_app = None
_cached_main_win = None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="微信批量发送 — Windows 版（pywinauto）")
    parser.add_argument("--dry", action="store_true", help="模拟运行")
    parser.add_argument("--call-single", action="store_true",
                        help="单次发送（由 cli.py call_sender 调用）")
    parser.add_argument("--target", type=str, help="发送目标")
    parser.add_argument("--msg-type", type=str, help="消息类型")
    parser.add_argument("--text", type=str, help="文字内容")
    parser.add_argument("--image-path", type=str, default="", help="图片路径")
    parser.add_argument("--verbose", action="store_true", help="详细日志")
    args = parser.parse_args()

    VERBOSE = args.verbose

    if args.call_single:
        if not args.target or not args.msg_type:
            print("[ERROR] --call-single 需要 --target 和 --msg-type", file=sys.stderr)
            sys.exit(1)
        try:
            call_send(args.target, args.msg_type, args.text or "", args.image_path or "")
            print("✅ 发送成功")
            sys.exit(0)
        except Exception as e:
            print(f"[ERROR] {e}", file=sys.stderr)
            sys.exit(1)

    cfg = load_cfg()
    xlsx_path = cfg.get("excel_path", "")
    send_interval = float(cfg.get("send_interval", 5))
    max_per_minute = int(cfg.get("max_per_minute", 8))
    dry_run = args.dry or cfg.get("dry_run", False)

    batch_send(
        dry_run=dry_run,
        send_interval=send_interval,
        max_per_minute=max_per_minute,
        xlsx_path=xlsx_path,
    )
